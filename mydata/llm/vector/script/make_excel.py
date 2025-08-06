#!/usr/bin/env python3
import os
import re
import json
from pathlib import Path
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed

from openpyxl import load_workbook
import xlrd  # for .xls

from uid_utils import generate_uid, get_relative_path  # ✅ 最終設計対応

TEXT_ROOT = Path("/mydata/llm/vector/db/text")
NAS_ROOT = Path("/mydata/nas")
TARGETS_JSONL = Path("/tmp/targets_text_excel.jsonl")

# ✅ MAX-2対応
MAX_WORKERS = max(1, os.cpu_count() - 2)

def clean_text(text: str) -> str:
    return re.sub(r"[\t\f\r]+", " ", text).strip()

def extract_text_from_excel(path: Path) -> str:
    if path.suffix.lower() == ".xls":
        wb = xlrd.open_workbook(str(path))
        output_lines = []
        for sheet in wb.sheets():
            output_lines.append(f"<<sheet:{sheet.name}>>")
            for row_idx in range(sheet.nrows):
                row = sheet.row_values(row_idx)
                line = " ".join(str(cell).strip() if cell else "" for cell in row)
                line = clean_text(line)
                if line:
                    output_lines.append(line)
            output_lines.append("")
        return "\n".join(output_lines).strip()
    else:
        wb = load_workbook(path, data_only=True)
        output_lines = []
        for sheet in wb.worksheets:
            output_lines.append(f"<<sheet:{sheet.title}>>")
            for row in sheet.iter_rows(values_only=True):
                if not row:
                    continue
                line = " ".join(str(cell).strip() if cell is not None else "" for cell in row)
                line = clean_text(line)
                if line:
                    output_lines.append(line)
            output_lines.append("")
        return "\n".join(output_lines).strip()

def save_text(filepath: Path, text: str):
    rel_path = filepath.relative_to(NAS_ROOT)
    out_path = TEXT_ROOT / rel_path.with_name(rel_path.name + ".txt")
    out_path.parent.mkdir(parents=True, exist_ok=True)

    uid = generate_uid(filepath)
    abs_path = out_path.resolve()
    rel_text_path = get_relative_path(out_path, TEXT_ROOT)
    ftype = "excel"
    stat = filepath.stat()
    mtime_iso = datetime.fromtimestamp(stat.st_mtime).isoformat()
    size = stat.st_size

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(f"[UID]: {uid}\n")
        f.write(f"[ABS_PATH]: {abs_path}\n")
        f.write(f"[REL_PATH]: {rel_text_path}\n")
        f.write(f"[TYPE]: {ftype}\n")
        f.write(f"[MTIME]: {mtime_iso}\n")
        f.write(f"[SIZE]: {size}\n")
        f.write("----------------------------------------\n")
        f.write(text)

def process_excel(filepath: Path):
    try:
        text = extract_text_from_excel(filepath)
        if text.strip():
            save_text(filepath, text)
            return f"[OK] {filepath}"
        else:
            return f"[WARN] 空データ: {filepath}"
    except Exception as e:
        return f"[ERROR] {filepath}\n{e}"

def main():
    if not TARGETS_JSONL.exists():
        print(f"[INFO] Excel対象なし: {TARGETS_JSONL}")
        return

    paths = []
    with TARGETS_JSONL.open("r", encoding="utf-8") as f:
        for line in f:
            try:
                entry = json.loads(line)
                filepath = NAS_ROOT / Path(entry["rel_path"])
                if filepath.exists() and filepath.suffix.lower() in [".xls", ".xlsx"]:
                    paths.append(filepath)
            except json.JSONDecodeError:
                continue

    if not paths:
        print("[INFO] 有効なExcelファイルがありません。")
        return

    print(f"[INFO] Excel処理開始: {len(paths)} 件")
    with ProcessPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(process_excel, p) for p in paths]
        for f in as_completed(futures):
            print(f.result())
    print("[DONE] Excel処理完了")

if __name__ == "__main__":
    main()










